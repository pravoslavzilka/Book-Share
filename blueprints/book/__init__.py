from flask import Blueprint, render_template, redirect, url_for, request, flash
from database import db_session
from models import Grade, Student, Book, BookType
from werkzeug.utils import secure_filename
from flask_login import login_required
import openpyxl


book_bp = Blueprint("book_bp",__name__,template_folder="templates")

ALLOWED_EXTENSIONS = {'xlsx','xlsm','xltx','xltm'}


@book_bp.route("/all/")
@login_required
def landing_page():
    books = Book.query.all()
    free_books = Book.query.filter(Book.student == None).count()
    rent_books = Book.query.filter(Book.student).count()
    total_books = Book.query.count()

    book_types = BookType.query.all()
    return render_template("book/landing_page.html",books=books,book_types=book_types,
                           ctbooks=total_books,cfbooks=free_books,crbooks=rent_books
                           )


@book_bp.route("/type/<bt_name>/")
@login_required
def book_type_page(bt_name):
    book_type = BookType.query.filter(BookType.name == bt_name).first()

    free_books = Book.query.filter(Book.student == None, Book.book_type == book_type).count()
    rent_books = Book.query.filter(Book.student, Book.book_type == book_type).count()
    total_books = Book.query.filter(Book.book_type == book_type).count()

    return render_template("book/book_type_page.html",book_type=book_type,
                           ctbooks=total_books, cfbooks=free_books, crbooks=rent_books
                           )


@book_bp.route("/total_list/")
@login_required
def list_of_books():
    books = Book.query.all()
    book_types = BookType.query.all()
    return render_template("book/book_list.html", books=books,book_types=book_types)


@book_bp.route("/list/<book_state>/")
@login_required
def books_state(book_state):
    book_types = BookType.query.all()
    if book_state == "free":
        books = Book.query.filter(Book.student_id == None).all()
    else:
        books = Book.query.filter(Book.student_id != None).all()

    return render_template("book/book_list.html",books=books,book_types=book_types)


@book_bp.route("/add_new/in/<bt_name>/",methods=["POST"])
@login_required
def add_book_with_type(bt_name):
    book_code = request.form["book_code"]
    book = Book.query.filter(Book.code == book_code).first()

    if book:
        flash("U??ebnica s t??mto k??dom u?? existuje", "danger")
        return redirect(url_for("book_bp.book_type_page",bt_name=bt_name))

    book_type = BookType.query.filter(BookType.name == bt_name).first()
    new_book = Book(int(book_code), book_type)

    db_session.add(new_book)
    try:
        db_session.commit()
    except OverflowError:
        flash("Neplatn?? k??d", "danger")
        return redirect(url_for("book_bp.book_type_page", bt_name=bt_name))

    flash(f"U??ebnica s k??dom {book_code} ??spe??ne pridan??","success")
    return redirect(url_for("book_bp.book_type_page", bt_name=bt_name))


@book_bp.route("/return/",methods=["GET"])
@login_required
def return_book_view():
    return render_template("book/return_book.html")


@book_bp.route("/return/",methods=["POST"])
@login_required
def return_book():
    code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == code).first()
    except OverflowError:
        flash("Neplatn?? k??d", "danger")
        return redirect(url_for("book_bp.return_book_view"))

    if book:
        if book.student:
            student = Student.query.filter(Student.id == book.student.id).first()
            student.books.remove(book)
            db_session.commit()
        flash(f"Kniha s k??dom {book.code} bola ??spe??ne vr??ten??","success")
        return redirect(url_for("book_bp.return_book_view"))

    flash("Neplatn?? k??d","danger")
    return redirect(url_for("book_bp.return_book_view"))


@book_bp.route("/return/from/list/",methods=["POST"])
@login_required
def return_book_from_list():
    code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == code).first()
    except OverflowError:
        flash("Neplatn?? k??d", "danger")
        return redirect(url_for("book_bp.list_of_books"))

    if book:
        if book.student:
            student = Student.query.filter(Student.id == book.student.id).first()
            student.books.remove(book)
            db_session.commit()
        flash(f"U??ebnica s k??dom {book.code} bola ??spe??ne vr??ten??", "success")
        return redirect(url_for("book_bp.list_of_books"))

    flash("Neplatn?? k??d", "danger")
    return redirect(url_for("book_bp.list_of_books"))


@book_bp.route("/return/with/type/",methods=["POST"])
@login_required
def return_book_with_type():
    code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == code).first()
    except OverflowError:
        flash("Neplatn?? k??d", "danger")
        return redirect(url_for("book_bp.book_type_page"))

    if book:
        if book.student:
            student = Student.query.filter(Student.id == book.student.id).first()
            student.books.remove(book)
            db_session.commit()
        flash(f"U??ebnica s k??dom {book.code} bola ??spe??ne vr??ten??", "success")
        return redirect(url_for("book_bp.book_type_page",bt_name=book.book_type.name))

    flash("Neplatn?? k??d", "danger")
    return redirect(url_for("book_bp.book_type_page",bt_name=book.book_type.name))


@book_bp.route("/delete/",methods=['POST'])
@login_required
def delete_book():
    book_code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == book_code).first()
    except OverflowError:
        flash("Neplatn?? k??d", "danger")
        return redirect(url_for("book_bp.list_of_books"))

    if book:
        book_type = BookType.query.filter(BookType.id == book.book_type_id).first()
        db_session.delete(book)
        db_session.commit()
        flash(f"Kniha s k??dom {book.code} bola ??spe??ne odstr??nen??","success")
        return redirect(url_for("book_bp.book_type_page",bt_name=book_type.name))

    flash("Neplatn?? k??d", "danger")
    return redirect(url_for("book_bp.list_of_books"))


@book_bp.route("/delete/from/list/",methods=['POST'])
@login_required
def delete_book_from_list():
    book_code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == book_code).first()
    except OverflowError:
        flash("Neplatn?? k??d","danger")
        return redirect(url_for("book_bp.list_of_books"))

    db_session.delete(book)
    db_session.commit()
    flash(f"Kniha s k??dom {book.code} bola ??spe??ne odstr??nen??","success")
    return redirect(url_for("book_bp.list_of_books"))


@book_bp.route("/add_type/",methods=["POST"])
@login_required
def add_book_type():
    req_bt = request.form["book_name"]
    req_ba = request.form["book_author"]

    try:
        book_type = BookType.query.filter(BookType.name == req_bt).first()
    except OverflowError:
        flash("Neplatn?? k??d","danger")
        return redirect(url_for("book_bp.landing_page"))

    if book_type:
        flash("Typ tejto u??ibnice u?? existuje","danger")
        return redirect(url_for("book_bp.landing_page"))

    new_bt = BookType(req_bt,req_ba)
    db_session.add(new_bt)
    try:
        db_session.commit()
    except OverflowError:
        flash("Neplatn?? ??daje", "danger")
        return redirect(url_for("book_bp.landing_page"))

    flash(f"Nov?? typ u??ebn??c {req_bt} bol ??spe??ne pridan??", "success")
    return redirect(url_for("book_bp.landing_page"))


@book_bp.route("/update/type/<int:type_id>/", methods=["POST"])
@login_required
def update_book_type(type_id):
    book_type = BookType.query.filter(BookType.id == type_id).first()

    if book_type:
        n_name = request.form["book_type_name"]
        n_author = request.form["book_type_author"]
        book_type.name = n_name
        book_type.author = n_author
        db_session.commit()
        flash(f"U??ebnica '{book_type.name}' bola aktualizovan??.","success")
        return redirect(url_for("book_bp.book_type_page", bt_name=book_type.name))
    else:
        return redirect(url_for("book_bp.landing_page"))


@book_bp.route("/delete/type/<int:type_id>/")
@login_required
def delete_type(type_id):

    try:
        book_type = BookType.query.filter(BookType.id == type_id).first()
    except OverflowError:
        flash("Neplatan?? k??d", "danger")
        return redirect(url_for("book_bp.landing_page"))

    count_of_books = 0
    if book_type:
        for index,book in enumerate(book_type.books):
            db_session.delete(book)
            count_of_books += 1

        db_session.delete(book_type)
        db_session.commit()
        flash(f"Typ u??ebnice '{book_type.name}' bol ??spe??ne vymazan?? a s n??m {count_of_books} u??ebn??c","success")
        return redirect(url_for("book_bp.landing_page"))

    flash("Neplatan?? k??d", "danger")
    return redirect(url_for("book_bp.landing_page"))


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@book_bp.route('/upload/source/excel/',methods=["POST"])
@login_required
def upload_file():

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active

        try:
            start_row = int(request.form["start-row"])
            col_name = int(request.form["col-name"])
            col_author = int(request.form["col-author"])
            col_code = int(request.form["col-code"])

            failed_items = 0
            new_items = 0

            for i in range(start_row-1,sheet_obj.max_row):
                name = sheet_obj.cell(row=i+1, column=col_name).value
                code = sheet_obj.cell(row=i+1, column=col_code).value

                book_type = BookType.query.filter(BookType.name == name).first()
                book = Book.query.filter(Book.code == code).first()
                if book_type:
                    if not book:
                        new_book = Book(code,book_type)
                        db_session.add(new_book)
                    failed_items += 1

                elif request.form.get("auto-fill") and col_name != "" and col_code != "" and col_author != "":
                    author = sheet_obj.cell(row=i+1, column=col_author).value
                    new_book_type = BookType(name,author)
                    db_session.add(new_book_type)
                    db_session.commit()

                    new_book = Book(code,new_book_type)
                    db_session.add(new_book)
                    db_session.commit()
                    new_items += 1

                else:
                    failed_items += 1

            db_session.commit()
        except:
            flash("Nastala chyba pri nahr??van??. Ujistite sa, ??i d??ta z tabu??ky nie su u?? v syst??me", "danger")
            return redirect(url_for("book_bp.landing_page"))

        if failed_items > 0 and request.form.get("auto-fill"):
            flash(f"{failed_items} chybn??ch/pr??zdnych/rovnak??ch riadkov nebolo nahrat??ch","info")

        if new_items > 0:
            flash(f"{new_items} nov??ch u??ebn??c bolo pridan??ch","info")

        flash("U??ebnice z execelu boli ??spe??ne pridan??","success")
        return redirect(url_for("book_bp.landing_page"))

    allow_f_string = ' / '.join(map(str, ALLOWED_EXTENSIONS))
    flash(f"S??bor nie je podporovan??. Typ s??boru mus?? by??: {allow_f_string}", "danger")
    return redirect(url_for("book_bp.landing_page"))
