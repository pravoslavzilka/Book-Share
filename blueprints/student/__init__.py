from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file
from database import db_session
from models import Grade, Student, Book
from werkzeug.utils import secure_filename
import openpyxl


student_bp = Blueprint("student_bp",__name__,template_folder="templates")

ALLOWED_EXTENSIONS = {'xlsx','xlsm','xltx','xltm'}
UPLOAD_FOLDER = '/path/files'


@student_bp.route("/all/")
def landing_page():
    grades = Grade.query.all()
    students = Student.query.all()
    return render_template("student/landing_page.html", grades=grades, students=students)


@student_bp.route("/grade/<grade>/")
def landing_page_grade(grade):
    grades = Grade.query.all()
    cer_grade = Grade.query.filter(Grade.name == grade).first()

    if cer_grade:
        flash(f"Zobrazujú sa študenti z ročníka {cer_grade.name}","info")
        return render_template("student/landing_page.html", grades=grades, students=cer_grade.students)
    flash(f"Trieda {grade} neexistuje","danger")
    return redirect(url_for("student_bp.landing_page"))


@student_bp.route("/new_student/",methods=["POST"])
def new_student():
    name = request.form["student_name"]
    code = request.form["student_code"]
    s_grade = request.form["grade"]
    grade = Grade.query.filter(Grade.name == s_grade).first()
    code_student = Student.query.filter(Student.code == code).first()
    if code_student:
        flash(f"Žiak s kódom {code} už existuje","danger")
        return redirect(url_for("student_bp.landing_page"))
    if grade:
        n_student = Student(name,grade,int(code))
        db_session.add(n_student)
        db_session.commit()
        flash(f"Študent {n_student.name} bol úspešne pridaný", "success")
        return redirect(url_for("student_bp.landing_page"))
    flash(f"Trieda {s_grade} neexistuje", "danger")
    return redirect(url_for("student_bp.landing_page"))


@student_bp.route("/move_all_students_up/")
def move_all_s_up():

    grades = ["Prima", "Sekunda", "Tercia", "Kvarta", "Kvinta", "Sexta", "Septima", "Oktava",
              "1.A","2.A","3.A","4.A",
              "1.B", "2.B", "3.B","4.B"
              ]
    final_grades = ["Oktava","4.A","4.B"]
    students = Student.query.all()

    for student in students:
        if student.grade.name in final_grades:
            db_session.delete(student)
            db_session.commit()
            continue
        index = grades.index(student.grade.name)
        new_grade = Grade.query.filter(Grade.name == grades[index+1]).first()
        student.grade = new_grade
        db_session.commit()

    flash("Všetci študenti prešli do vyšieho ročníka", "success")
    return redirect(url_for("student_bp.landing_page"))


@student_bp.route("/move_all_students_down/")
def move_all_s_down():

    grades = ["Prima", "Sekunda", "Tercia", "Kvarta", "Kvinta", "Sexta", "Septima", "Oktava",
              "1.A","2.A","3.A","4.A",
              "1.B", "2.B", "3.B","4.B"
              ]
    final_grades = ["Prima","1.A","1.B"]
    students = Student.query.all()

    for student in students:
        if student.grade.name in final_grades:
            db_session.delete(student)
            db_session.commit()
            continue
        index = grades.index(student.grade.name)
        new_grade = Grade.query.filter(Grade.name == grades[index-1]).first()
        student.grade = new_grade
        db_session.commit()

    flash("Všetci študenti prešli do nižšieho ročníka", "success")
    return redirect(url_for("student_bp.landing_page"))


@student_bp.route("/view/<int:student_id>/")
def view_student(student_id):
    try:
        student = Student.query.filter(Student.id == student_id).first()
    except OverflowError:
        flash("Neplatný kód študenta", "danger")
        return redirect(url_for("student_bp.search_student"))

    if student:
        grades = Grade.query.all()
        return render_template("student/student_page.html",student=student, grades=grades)
    flash("Neplatný kód študenta","danger")
    return redirect(url_for("student_bp.search_student"))


@student_bp.route("/change_grade/for/<int:student_id>/",methods=["POST"])
def change_student(student_id):
    new_name = request.form["student_name"]
    new_code = request.form["student_code"]
    new_grade = Grade.query.filter(Grade.name == request.form["student_grade"]).first()

    student = Student.query.filter(Student.id == student_id).first()
    if not student:
        flash("Neplatný kód študenta","danger")
        return redirect(url_for("student_bp.landing_page"))
    code_student = Student.query.filter(Student.code == new_code).first()

    if code_student:
        flash("Študent s takýmto kódom už existuje","danger")
        return redirect(url_for("student_bp.view_student",student_id=student_id))

    student.name = new_name
    student.code = new_code
    student.grade = new_grade
    db_session.commit()

    flash(f"Údaje študenta {student.name} boli úspešne zmenené", "success")
    return redirect(url_for("student_bp.view_student",student_id=student_id))


@student_bp.route("/<int:student_id>/rent_book/",methods=["POST"])
def rent_book(student_id):
    try:
        book = Book.query.filter(Book.code == int(request.form["code"])).first()
    except OverflowError:
        flash("Neplatný kód učebnice","danger")
        return redirect(url_for("student_bp.view_student", student_id=student_id))

    if book:
        if book.student:
            flash("Učebnica s týmto kódom je už požičaná","danger")
        else:
            student = Student.query.filter(Student.id == student_id).first()
            if student:
                student.books.append(book)
                db_session.commit()
                flash("Učebnica bola pridaná", "success")
            else:
                return redirect(url_for("student_bp.search_student"))
    else:
        flash("Neplatný kód","danger")
    return redirect(url_for("student_bp.view_student",student_id=student_id))


@student_bp.route("/<int:student_id>/return_book/<int:book_id>/")
def return_book(student_id,book_id):

    book = Book.query.filter(Book.id == book_id).first()
    student = Student.query.filter(Student.id == student_id).first()

    if book and student:
        student.books.remove(book)
        db_session.commit()
        flash("Učebnica bola vrátená","success")

        return redirect(url_for("student_bp.view_student",student_id=student_id))
    else:
        return redirect(url_for("student_bp.landing_page"))


@student_bp.route("/<int:student_id>/return_all/")
def return_all(student_id):
    student = Student.query.filter(Student.id == student_id).first()
    if not student:
        return redirect(url_for("student_bp.landing_page"))

    student.books.clear()
    db_session.commit()
    flash(f"Všetky učebnice žiaka {student.name} boli úspešne vrátené","success")

    return redirect(url_for("student_bp.view_student",student_id=student_id))


@student_bp.route("/delete/<int:student_id>/")
def delete_student(student_id):
    student = Student.query.filter(Student.id == student_id).first()
    if not student:
        return redirect(url_for("student_bp.landing_page"))
    db_session.delete(student)
    db_session.commit()
    flash(f"Študent {student.name} bol úspešne vymazaný","success")
    return redirect(url_for("student_bp.landing_page"))


@student_bp.route("/search_student/")
def search_student():
    return render_template("student/search_student.html")


@student_bp.route("/search_student/",methods=["POST"])
def search_student2():
    student_code = request.form["student_code"]
    if 1 < len(student_code) < 25:
        student = Student.query.filter(Student.code == student_code).first()
        if student:
            return redirect(url_for("student_bp.view_student",student_id=student.id))

    flash("Neplatný kód","danger")
    return render_template("student/search_student.html")


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@student_bp.route('/upload/source/excel/',methods=["POST"])
def upload_file():

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active

        try:
            for i in range(1,sheet_obj.max_row):
                name = sheet_obj.cell(row=i+1, column=2).value
                code = sheet_obj.cell(row=i+1, column=3).value
                grade_name = sheet_obj.cell(row=i+1, column=4).value
                grade = Grade.query.filter(Grade.name == grade_name).first()

                s = Student(name,grade,code)
                db_session.add(s)
                db_session.commit()
        except:
            flash("Nastala chyba pri nahrávaní. Ujistite sa, či študenti z tabuľky nie su už v systéme", "danger")
            return redirect(url_for("student_bp.landing_page"))

        flash("Študenti z execelu boli úspešne pridaný","success")
        return redirect(url_for("student_bp.landing_page"))

    allow_f_string = ' / '.join(map(str, ALLOWED_EXTENSIONS))
    flash(f"Súbor nie je podporovaný. Typ súboru musí byť: {allow_f_string}","danger")
    return redirect(url_for("student_bp.landing_page"))

